# excel_export.py
# Module d'export Excel multi-feuilles — DGID Suivi
# v1.0.0 — Génération en mémoire, style DGID, feuilles dynamiques

import io
import os
import duckdb
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# =============================================================================
# CONSTANTES DGID
# =============================================================================
GOLD_HEX = "D4AF37"
BROWN_HEX = "4A3525"
CREAM_HEX = "FAF7F2"
WHITE_HEX = "FFFFFF"
TEXT_MAIN_HEX = "1E1E1E"
TEXT_MUTED_HEX = "8A8A8A"

# =============================================================================
# STYLES RÉUTILISABLES
# =============================================================================
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=BROWN_HEX)
FONT_SUBTITLE = Font(name="Calibri", size=10, color=TEXT_MUTED_HEX, italic=True)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=BROWN_HEX)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
FONT_KPI_LABEL = Font(name="Calibri", size=10, color=TEXT_MUTED_HEX)
FONT_KPI_VALUE = Font(name="Calibri", size=22, bold=True, color=BROWN_HEX)
FONT_NORMAL = Font(name="Calibri", size=11, color=TEXT_MAIN_HEX)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=BROWN_HEX)

FILL_HEADER = PatternFill(start_color=GOLD_HEX, end_color=GOLD_HEX, fill_type="solid")
FILL_CREAM = PatternFill(start_color=CREAM_HEX, end_color=CREAM_HEX, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE_HEX, end_color=WHITE_HEX, fill_type="solid")
FILL_LIGHT_GOLD = PatternFill(start_color="F5E6C8", end_color="F5E6C8", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="E8E8E8"),
    right=Side(style="thin", color="E8E8E8"),
    top=Side(style="thin", color="E8E8E8"),
    bottom=Side(style="thin", color="E8E8E8"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =============================================================================
# HELPERS SQL & DONNÉES
# =============================================================================
def _build_where(filters: dict) -> tuple[str, list]:
    clauses = []
    params = []
    if filters.get("start_date") and filters.get("end_date"):
        clauses.append("registration_date::DATE BETWEEN ?::DATE AND ?::DATE")
        params.extend([filters["start_date"], filters["end_date"]])
    if filters.get("service") and filters["service"] != "Tous":
        clauses.append("service_origine = ?")
        params.append(filters["service"])
    if filters.get("csf") and filters["csf"] != "Tous":
        clauses.append("csf_geographique = ?")
        params.append(filters["csf"])
    if filters.get("type") and filters["type"] != "Tous":
        clauses.append("motif_enregistrement = ?")
        params.append(filters["type"])
    where_sql = " AND ".join(clauses) if clauses else "1=1"
    return where_sql, params


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# =============================================================================
# STYLING HELPERS
# =============================================================================
def _apply_header_style(ws, row_idx, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _apply_row_style(ws, row_idx, col_count, bold=False, fill=None):
    font = FONT_BOLD if bold else FONT_NORMAL
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = font
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
        if fill:
            cell.fill = fill


def _auto_adjust_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 55)
        ws.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# FEUILLE D'AGRÉGATION PAR OPÉRATEUR
# =============================================================================
def _write_aggregation_sheet(ws, df_sub, title):
    """Écrit une feuille d'agrégation par opérateur avec style DGID."""
    # Titre
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # Sous-titre
    ws.merge_cells("A2:G2")
    nb_lignes = len(df_sub)
    ws["A2"] = f"{nb_lignes} enregistrements bruts | Export DGID v1.0.0"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 18

    # Headers
    agg_headers = [
        "Opérateur",
        "Dossiers Traités",
        "Pièces Indexées",
        "En Attente Num.",
        "Taux Restitution",
        "Durée Moyenne (h)",
        "Dossiers / Jour",
    ]
    for j, h in enumerate(agg_headers, 1):
        ws.cell(row=4, column=j, value=h)
    _apply_header_style(ws, 4, len(agg_headers))

    # Données
    row = 5
    operators = df_sub.groupby("operateur")
    for op_name, df_op in operators:
        if pd.isna(op_name) or str(op_name).strip() == "":
            op_name = "Operateur Inconnu"

        nb_dossiers = df_op["numero_dossier"].nunique()
        nb_pieces = df_op[df_op["date_indexation"].notna()]["count_item"].sum()
        if pd.isna(nb_pieces):
            nb_pieces = 0

        attente = df_op[
            (df_op["is_for_scan"] == True) & (df_op["date_indexation"].isna())
        ]["numero_dossier"].nunique()

        rest = df_op[df_op["date_retour"].notna()]["numero_dossier"].nunique()
        taux_rest = round((rest / nb_dossiers * 100), 1) if nb_dossiers > 0 else 0

        # Durée moyenne
        df_duree = df_op[
            (df_op["date_indexation"].notna()) & (df_op["registration_date"].notna())
        ].copy()
        duree = 0.0
        if not df_duree.empty:
            df_duree["duree_h"] = (
                df_duree["date_indexation"] - df_duree["registration_date"]
            ).dt.total_seconds() / 3600
            duree = round(df_duree["duree_h"].mean(), 1)

        # Dossiers par jour
        nb_j = df_op[df_op["registration_date"].notna()]["registration_date"].dt.date.nunique()
        dpj = round(nb_dossiers / max(nb_j, 1), 1)

        vals = [op_name, nb_dossiers, int(nb_pieces), attente, f"{taux_rest}%", duree, dpj]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
        row += 1

    _auto_adjust_columns(ws)


# =============================================================================
# GÉNÉRATEUR PRINCIPAL
# =============================================================================
def generate_excel_report(duckdb_path: str, filters: dict, objectif: int) -> bytes:
    """Génère le classeur Excel complet en mémoire et retourne les bytes."""
    if not os.path.exists(duckdb_path):
        raise FileNotFoundError(f"Base DuckDB introuvable : {duckdb_path}")

    con = duckdb.connect(duckdb_path, read_only=True)
    try:
        where_sql, params = _build_where(filters)

        # ---------------------------------------------------------------------
        # 1. RÉCUPÉRATION DES DONNÉES FILTRÉES
        # ---------------------------------------------------------------------
        df = con.execute(
            f"SELECT * FROM fact_suivi_global WHERE {where_sql}", params
        ).fetchdf()

        if df.empty:
            raise ValueError("Aucune donnée à exporter pour les filtres sélectionnés.")

        # Normalisation des dates
        for col in ["registration_date", "date_retour", "date_indexation", "date_maj"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        # ---------------------------------------------------------------------
        # 2. CALCUL DES KPIs GLOBAUX
        # ---------------------------------------------------------------------
        total_dossiers = df["numero_dossier"].nunique()
        pieces_indexees = df[df["date_indexation"].notna()]["count_item"].sum()
        if pd.isna(pieces_indexees):
            pieces_indexees = 0

        dossiers_restitués = df[df["date_retour"].notna()]["numero_dossier"].nunique()
        taux_restitution = (
            round((dossiers_restitués / total_dossiers * 100), 1)
            if total_dossiers > 0
            else 0
        )

        attente_num = df[
            (df["is_for_scan"] == True) & (df["date_indexation"].isna())
        ]["numero_dossier"].nunique()

        # Durée moyenne
        df_duree = df[
            (df["date_indexation"].notna()) & (df["registration_date"].notna())
        ].copy()
        duree_moyenne = 0.0
        if not df_duree.empty:
            df_duree["duree_h"] = (
                df_duree["date_indexation"] - df_duree["registration_date"]
            ).dt.total_seconds() / 3600
            duree_moyenne = round(df_duree["duree_h"].mean(), 1)

        # Dossiers / jour
        df_jours = df[df["registration_date"].notna()].copy()
        nb_jours = df_jours["registration_date"].dt.date.nunique()
        dossiers_jour = round(total_dossiers / max(nb_jours, 1), 1)

        # Top agent
        top_serie = (
            df.groupby("operateur")["numero_dossier"]
            .nunique()
            .sort_values(ascending=False)
            .head(1)
        )
        top_agent_name = str(top_serie.index[0]) if not top_serie.empty else "—"
        top_agent_count = int(top_serie.iloc[0]) if not top_serie.empty else 0

        # ---------------------------------------------------------------------
        # 3. CRÉATION DU CLASSEUR EXCEL
        # ---------------------------------------------------------------------
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine="openpyxl")
        workbook = writer.book

        # =================================================================
        # FEUILLE 1 : RÉSUMÉ GLOBAL
        # =================================================================
        ws_resume = workbook.create_sheet("Résumé Global", 0)

        # Titre principal
        ws_resume.merge_cells("A1:G1")
        ws_resume["A1"] = (
            "RAPPORT DE SUIVI GLOBAL - DIRECTION GENERALE DES IMPOTS ET DES DOMAINES (DGID)"
        )
        ws_resume["A1"].font = FONT_TITLE
        ws_resume["A1"].alignment = ALIGN_CENTER
        ws_resume.row_dimensions[1].height = 30

        # Sous-titre filtres
        filter_text = f"Généré le : {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        filter_text += f"CSF : {filters.get('csf', 'Tous')}  |  "
        filter_text += f"Bureaux : {filters.get('service', 'Tous')}  |  "
        filter_text += f"Période : {filters.get('start_date', '—')} au {filters.get('end_date', '—')}"
        ws_resume.merge_cells("A2:G2")
        ws_resume["A2"] = filter_text
        ws_resume["A2"].font = FONT_SUBTITLE
        ws_resume["A2"].alignment = ALIGN_CENTER
        ws_resume.row_dimensions[2].height = 20

        # --- KPIs (ligne 4-6) ---
        kpi_data = [
            ("TOTAL DOSSIERS", str(total_dossiers)),
            ("PIÈCES INDEXÉES", str(int(pieces_indexees))),
            ("TAUX RESTITUTION", f"{taux_restitution} %"),
            ("ATTENTE NUMÉRISATION", str(attente_num)),
            ("DURÉE MOYENNE", f"{duree_moyenne} h"),
            ("TOP AGENT", f"{top_agent_name} ({top_agent_count})"),
        ]

        for i, (label, value) in enumerate(kpi_data):
            col_start = i + 1
            # Label
            cell_lbl = ws_resume.cell(row=4, column=col_start)
            cell_lbl.value = label
            cell_lbl.font = FONT_KPI_LABEL
            cell_lbl.alignment = ALIGN_CENTER
            # Value
            cell_val = ws_resume.cell(row=5, column=col_start)
            cell_val.value = value
            cell_val.font = FONT_KPI_VALUE
            cell_val.alignment = ALIGN_CENTER
            # Fond & bordure
            for r in range(4, 6):
                c = ws_resume.cell(row=r, column=col_start)
                c.fill = FILL_CREAM
                c.border = BORDER_THIN

        ws_resume.row_dimensions[4].height = 18
        ws_resume.row_dimensions[5].height = 32

        # --- Synthèse par Bureau ---
        row = 8
        ws_resume.merge_cells(f"A{row}:G{row}")
        ws_resume.cell(row=row, column=1).value = "SYNTHÈSE DE PRODUCTION PAR BUREAU"
        ws_resume.cell(row=row, column=1).font = FONT_SECTION
        ws_resume.cell(row=row, column=1).alignment = ALIGN_LEFT
        row += 1

        headers_bureau = [
            "Bureau",
            "Dossiers Traités",
            "Pièces Indexées",
            "En Attente Num.",
            "Objectif Période",
            "% Réalisation",
            "Taux Restitution",
        ]
        for j, h in enumerate(headers_bureau, 1):
            ws_resume.cell(row=row, column=j, value=h)
        _apply_header_style(ws_resume, row, len(headers_bureau))
        row += 1

        for bureau in sorted(df["service_origine"].dropna().unique()):
            df_b = df[df["service_origine"] == bureau]
            nb_dossiers = df_b["numero_dossier"].nunique()
            nb_pieces = df_b[df_b["date_indexation"].notna()]["count_item"].sum()
            if pd.isna(nb_pieces):
                nb_pieces = 0
            attente_b = df_b[
                (df_b["is_for_scan"] == True) & (df_b["date_indexation"].isna())
            ]["numero_dossier"].nunique()
            rest_b = df_b[df_b["date_retour"].notna()]["numero_dossier"].nunique()
            taux_rest_b = round((rest_b / nb_dossiers * 100), 1) if nb_dossiers > 0 else 0
            pct_real = round((nb_dossiers / max(objectif, 1)) * 100, 1)

            vals = [
                bureau,
                nb_dossiers,
                int(nb_pieces),
                attente_b,
                objectif,
                f"{pct_real} %",
                f"{taux_rest_b} %",
            ]
            for j, v in enumerate(vals, 1):
                cell = ws_resume.cell(row=row, column=j, value=v)
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN
            row += 1

        # Ligne Total
        vals_total = [
            "TOTAL",
            total_dossiers,
            int(pieces_indexees),
            attente_num,
            objectif,
            f"{round((total_dossiers / max(objectif, 1)) * 100, 1)} %",
            f"{taux_restitution} %",
        ]
        for j, v in enumerate(vals_total, 1):
            cell = ws_resume.cell(row=row, column=j, value=v)
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
            cell.fill = FILL_LIGHT_GOLD

        # --- Ventilation par CSF ---
        row += 3
        ws_resume.merge_cells(f"A{row}:F{row}")
        ws_resume.cell(row=row, column=1).value = (
            "VENTILATION DE LA PRODUCTION PAR CENTRE DES SERVICES FISCAUX (CSF)"
        )
        ws_resume.cell(row=row, column=1).font = FONT_SECTION
        ws_resume.cell(row=row, column=1).alignment = ALIGN_LEFT
        row += 1

        headers_csf = [
            "CSF Géographique",
            "Bureau",
            "Dossiers Traités",
            "Pièces Indexées",
            "Taux Restitution",
            "Dossiers / Jour",
        ]
        for j, h in enumerate(headers_csf, 1):
            ws_resume.cell(row=row, column=j, value=h)
        _apply_header_style(ws_resume, row, len(headers_csf))
        row += 1

        for csf in sorted(df["csf_geographique"].dropna().unique()):
            for bureau in sorted(
                df[df["csf_geographique"] == csf]["service_origine"].dropna().unique()
            ):
                df_cb = df[
                    (df["csf_geographique"] == csf) & (df["service_origine"] == bureau)
                ]
                nb_d = df_cb["numero_dossier"].nunique()
                nb_p = df_cb[df_cb["date_indexation"].notna()]["count_item"].sum()
                if pd.isna(nb_p):
                    nb_p = 0
                rest = df_cb[df_cb["date_retour"].notna()]["numero_dossier"].nunique()
                taux_r = round((rest / nb_d * 100), 1) if nb_d > 0 else 0
                nb_j = df_cb[df_cb["registration_date"].notna()][
                    "registration_date"
                ].dt.date.nunique()
                dpj = round(nb_d / max(nb_j, 1), 1)

                vals = [csf, bureau, nb_d, int(nb_p), f"{taux_r} %", dpj]
                for j, v in enumerate(vals, 1):
                    cell = ws_resume.cell(row=row, column=j, value=v)
                    cell.font = FONT_NORMAL
                    cell.alignment = ALIGN_CENTER
                    cell.border = BORDER_THIN
                row += 1

        _auto_adjust_columns(ws_resume)

        # =================================================================
        # FEUILLES PAR CSF (si pertinent)
        # =================================================================
        csf_list = sorted(df["csf_geographique"].dropna().unique())
        for csf in csf_list:
            df_csf = df[df["csf_geographique"] == csf]
            # Limite Excel : 31 caractères par nom de feuille
            sheet_name = csf[:31]
            ws = workbook.create_sheet(sheet_name)
            _write_aggregation_sheet(ws, df_csf, f"PRODUCTION - CSF {csf.upper()}")

        # =================================================================
        # FEUILLES PAR BUREAU "All X" (si plus d'un CSF)
        # =================================================================
        bureau_list = sorted(df["service_origine"].dropna().unique())
        if len(csf_list) > 1:
            for bureau in bureau_list:
                df_bureau = df[df["service_origine"] == bureau]
                sheet_name = f"All {bureau}"[:31]
                ws = workbook.create_sheet(sheet_name)
                _write_aggregation_sheet(
                    ws, df_bureau, f"PRODUCTION - ALL {bureau.upper()}"
                )

        # =================================================================
        # FEUILLES PAR COMBINAISON CSF × BUREAU
        # =================================================================
        # On évite le doublon si un seul CSF et un seul bureau (déjà couvert)
        for csf in csf_list:
            for bureau in bureau_list:
                df_combo = df[
                    (df["csf_geographique"] == csf)
                    & (df["service_origine"] == bureau)
                ]
                if not df_combo.empty:
                    if len(csf_list) == 1 and len(bureau_list) == 1:
                        continue  # Déjà traité
                    sheet_name = f"{bureau} {csf}"[:31]
                    ws = workbook.create_sheet(sheet_name)
                    _write_aggregation_sheet(
                        ws, df_combo, f"PRODUCTION - {bureau.upper()} {csf.upper()}"
                    )

        # =================================================================
        # FEUILLE DÉTAIL BRUT (500 derniers dossiers)
        # =================================================================
        ws_detail = workbook.create_sheet("Détail Brut")
        df_detail = df.sort_values("registration_date", ascending=False).head(500)

        # Titre
        ws_detail.merge_cells("A1:J1")
        ws_detail["A1"] = "REGISTRE DÉTAILLÉ DES DOSSIERS EXTRAITS"
        ws_detail["A1"].font = FONT_TITLE
        ws_detail["A1"].alignment = ALIGN_CENTER
        ws_detail.row_dimensions[1].height = 28

        # Sous-titre
        ws_detail.merge_cells("A2:J2")
        ws_detail["A2"] = f"500 derniers enregistrements | Export DGID v1.0.0"
        ws_detail["A2"].font = FONT_SUBTITLE
        ws_detail["A2"].alignment = ALIGN_CENTER

        # Headers
        detail_headers = [
            "N° Dossier",
            "CSF",
            "Bureau",
            "Opérateur",
            "Date Enregistrement",
            "Date Retour",
            "Date Indexation",
            "Statut",
            "Pièces Indexées",
            "Anomalie Chrono",
        ]
        for j, h in enumerate(detail_headers, 1):
            ws_detail.cell(row=4, column=j, value=h)
        _apply_header_style(ws_detail, 4, len(detail_headers))

        # Données
        for idx, row_data in df_detail.iterrows():
            r = ws_detail.max_row + 1
            if pd.notna(row_data.get("date_indexation")):
                statut = "Indexé"
            elif pd.notna(row_data.get("registration_date")):
                statut = "En cours"
            else:
                statut = "—"

            anomalie = (
                "Oui"
                if row_data.get("anomalie_chronologie") == True
                else "Non"
            )

            vals = [
                row_data.get("numero_dossier", "—"),
                row_data.get("csf_geographique", "—"),
                row_data.get("service_origine", "—"),
                row_data.get("operateur", "—"),
                row_data["registration_date"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("registration_date"))
                else "—",
                row_data["date_retour"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("date_retour"))
                else "—",
                row_data["date_indexation"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("date_indexation"))
                else "—",
                statut,
                int(row_data.get("count_item", 0))
                if pd.notna(row_data.get("count_item"))
                else 0,
                anomalie,
            ]
            for j, v in enumerate(vals, 1):
                cell = ws_detail.cell(row=r, column=j, value=v)
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN

        _auto_adjust_columns(ws_detail)

        # ---------------------------------------------------------------------
        # 4. FINALISATION
        # ---------------------------------------------------------------------
        writer.close()
        output.seek(0)
        return output.read()

    finally:
        con.close()