# suivi_app.py
# Dashboard de Suivi DGID — Reflex
# v3.0.0 — FIX : Validation dates, logique Entrée/Sortie via record_type+output_id, Complexité onomastique

import reflex as rx
from datetime import datetime, timedelta
import math
import asyncio
import duckdb
import os
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONSTANTES DE STYLE
# =============================================================================
GOLD = "#D4AF37"
BROWN = "#4A3525"
CREAM = "#FAF7F2"
WHITE = "#FFFFFF"
TEXT_MAIN = "#1E1E1E"
TEXT_MUTED = "#8A8A8A"
BORDER_HDR = "#E8E8E8"
BORDER_CARD = "#F0F0F0"
ICON_BG = "#F5F0EB"
CREAM_CARD = "#FEF7E0"
BROWN_DARK = "#4A3525"
GREEN_LIGHT = "#E6F4EA"
GREEN_TEXT = "#1E8E3E"
RED_LIGHT = "#FCE8E6"
RED_TEXT = "#D93025"
CADASTRE_BG = "#F5E6C8"
CADASTRE_TEXT = "#8B6914"
CONSERVATION_BG = "#E8E8E8"
CONSERVATION_TEXT = "#666666"
DOMAINES_BG = "#E6D5C3"
DOMAINES_TEXT = "#6B4423"
SHADOW = "0 2px 12px rgba(0,0,0,0.04)"

DUCKDB_PATH = os.getenv("DUCKDB_PATH", "/app/duckdb_store/analytics.duckdb")

# =============================================================================
# CONSTANTES MÉTIER — UUID des types d'enregistrement
# =============================================================================
UUID_ENTREE = "1647d4a7-785a-4419-a248-d62259683da5"
UUID_SORTIE = "8fba2098-1356-4f34-9d29-68e5446f49a5"

# =============================================================================
# CONSTANTES EXCEL
# =============================================================================
GOLD_HEX = "D4AF37"
BROWN_HEX = "4A3525"
CREAM_HEX = "FAF7F2"
WHITE_HEX = "FFFFFF"
TEXT_MAIN_HEX = "1E1E1E"
TEXT_MUTED_HEX = "8A8A8A"

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=BROWN_HEX)
FONT_SUBTITLE = Font(name="Calibri", size=10, color=TEXT_MUTED_HEX, italic=True)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=BROWN_HEX)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE_HEX)
FONT_KPI_LABEL = Font(name="Calibri", size=10, color=TEXT_MUTED_HEX)
FONT_KPI_VALUE = Font(name="Calibri", size=22, bold=True, color=BROWN_HEX)
FONT_NORMAL = Font(name="Calibri", size=11, color=TEXT_MAIN_HEX)
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=BROWN_HEX)

FILL_HEADER = PatternFill(start_color=GOLD_HEX, end_color=GOLD_HEX, fill_type="solid")
FILL_CREAM = PatternFill(start_color=CREAM_HEX, end_color=CREAM_HEX, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE_HEX, end_color=WHITE_HEX, fill_type="solid")
FILL_LIGHT_GOLD = PatternFill(start_color="F5E6C8", end_color="F5E6C8", fill_type="solid")
FILL_GREEN_LIGHT = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
FILL_RED_LIGHT = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style="thin", color="E8E8E8"),
    right=Side(style="thin", color="E8E8E8"),
    top=Side(style="thin", color="E8E8E8"),
    bottom=Side(style="thin", color="E8E8E8"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# =============================================================================
# HELPERS
# =============================================================================
def _build_where(filters: dict) -> tuple[str, list]:
    """Construit la clause WHERE."""
    clauses = []
    params = []
    
    if filters.get("start_date") and filters.get("end_date"):
        clauses.append("registration_date::DATE BETWEEN ?::DATE AND ?::DATE")
        params.extend([filters["start_date"], filters["end_date"]])
    if filters.get("service") and filters["service"] != "Tous":
        clauses.append("service_origine = ?")
        params.append(filters["service"])
    if filters.get("csf") and filters["csf"] != "Tous":
        clauses.append("csf_geographique = ?")
        params.append(filters["csf"])
    if filters.get("type") and filters["type"] != "Tous":
        clauses.append("vrai_type = ?")
        params.append(filters["type"])
    where_sql = " AND ".join(clauses) if clauses else "1=1"
    return where_sql, params


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _validate_date_range(start_date: str, end_date: str) -> tuple[bool, str]:
    """Valide que la plage de dates est cohérente."""
    if not start_date or not end_date:
        return True, ""
    try:
        sd = datetime.strptime(start_date, "%Y-%m-%d")
        ed = datetime.strptime(end_date, "%Y-%m-%d")
        if ed < sd:
            return False, "La date de fin doit être postérieure à la date de début."
        # Limite max : pas de date future
        if ed > datetime.now() + timedelta(days=1):
            return False, "La date de fin ne peut pas être dans le futur."
        return True, ""
    except ValueError:
        return False, "Format de date invalide."


# =============================================================================
# HELPERS EXCEL STYLING
# =============================================================================
def _apply_header_style(ws, row_idx, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def _auto_adjust_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 55)
        ws.column_dimensions[column_letter].width = adjusted_width


def _write_aggregation_sheet(ws, df_sub, title):
    """Écrit une feuille d'agrégation par opérateur avec style DGID + complexité."""
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    nb_lignes = len(df_sub)
    ws["A2"] = f"{nb_lignes} enregistrements bruts"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 18

    # v3.0 : Ajout colonne Complexité Moyenne
    agg_headers = [
        "Opérateur",
        "Dossiers Traités",
        "Pièces Indexées",
        "En Attente Num.",
        "Taux Restitution",
        "Durée Moyenne Sortie (h)",
        "Dossiers / Jour",
        "Complexité Moy.",
    ]
    for j, h in enumerate(agg_headers, 1):
        ws.cell(row=4, column=j, value=h)
    _apply_header_style(ws, 4, len(agg_headers))

    row = 5
    operators = df_sub.groupby("operateur")
    for op_name, df_op in operators:
        if pd.isna(op_name) or str(op_name).strip() == "":
            op_name = "Operateur Inconnu"

        nb_dossiers = df_op["numero_dossier"].nunique()
        # v3.1 : count_item par dossier, pas par mouvement -> dédup avant sum
        df_idx_op = df_op[df_op["date_indexation"].notna()].drop_duplicates(subset=["numero_dossier"])
        nb_pieces = df_idx_op["count_item"].sum()
        if pd.isna(nb_pieces):
            nb_pieces = 0

        attente = df_op[
            (df_op["is_for_scan"] == True) & (df_op["date_indexation"].isna())
        ]["numero_dossier"].nunique()

        rest = df_op[df_op["date_retour"].notna()]["numero_dossier"].nunique()
        taux_rest = round((rest / nb_dossiers * 100), 1) if nb_dossiers > 0 else 0

        duree = 0.0
        if "duree_numerisation_h" in df_op.columns:
            vals = df_op["duree_numerisation_h"].dropna()
            if not vals.empty:
                duree = round(vals.mean(), 1)
        else:
            df_duree = df_op[
                (df_op["date_retour"].notna()) & (df_op["date_indexation"].notna())
            ].copy()
            if not df_duree.empty:
                df_duree["duree_h"] = (
                    df_duree["date_retour"] - df_duree["date_indexation"]
                ).dt.total_seconds() / 3600
                df_duree = df_duree[df_duree["duree_h"] >= 0]
                if not df_duree.empty:
                    duree = round(df_duree["duree_h"].mean(), 1)

        nb_j = df_op[df_op["registration_date"].notna()]["registration_date"].dt.date.nunique()
        dpj = round(nb_dossiers / max(nb_j, 1), 1)

        # v3.0 : Complexité moyenne
        complexite_moy = 0.0
        if "nb_proprietaires_uniques" in df_op.columns:
            comp_vals = df_op["nb_proprietaires_uniques"].dropna()
            if not comp_vals.empty:
                complexite_moy = round(comp_vals.mean(), 1)

        vals = [op_name, nb_dossiers, int(nb_pieces), attente, f"{taux_rest}%", duree, dpj, complexite_moy]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
        row += 1

    _auto_adjust_columns(ws)


# =============================================================================
# GÉNÉRATEUR EXCEL PRINCIPAL
# =============================================================================
def generate_excel_report(duckdb_path: str, filters: dict, objectif: int) -> bytes:
    """Génère le classeur Excel complet en mémoire et retourne les bytes."""
    if not os.path.exists(duckdb_path):
        raise FileNotFoundError(f"Base DuckDB introuvable : {duckdb_path}")

    con = duckdb.connect(duckdb_path, read_only=True)
    try:
        where_sql, params = _build_where(filters)

        df = con.execute(
            f"SELECT * FROM fact_suivi_global WHERE {where_sql}", params
        ).fetchdf()

        if df.empty:
            raise ValueError("Aucune donnée à exporter pour les filtres sélectionnés.")

        for col in ["registration_date", "date_retour", "date_indexation", "date_maj"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")

        total_dossiers = df["numero_dossier"].nunique()
        # v3.1 : count_item par dossier, pas par mouvement -> dédup avant sum
        pieces_indexees = df[df["date_indexation"].notna()].drop_duplicates(subset=["numero_dossier"])["count_item"].sum()
        if pd.isna(pieces_indexees):
            pieces_indexees = 0

        dossiers_restitués = df[df["date_retour"].notna()]["numero_dossier"].nunique()
        taux_restitution = round((dossiers_restitués / total_dossiers * 100), 1) if total_dossiers > 0 else 0

        attente_num = df[
            (df["is_for_scan"] == True) & (df["date_indexation"].isna())
        ]["numero_dossier"].nunique()

        duree_moyenne = 0.0
        if "duree_numerisation_h" in df.columns:
            vals = df["duree_numerisation_h"].dropna()
            if not vals.empty:
                duree_moyenne = round(vals.mean(), 1)
        else:
            df_duree = df[
                (df["date_retour"].notna()) & (df["date_indexation"].notna())
            ].copy()
            if not df_duree.empty:
                df_duree["duree_h"] = (
                    df_duree["date_retour"] - df_duree["date_indexation"]
                ).dt.total_seconds() / 3600
                df_duree = df_duree[df_duree["duree_h"] >= 0]
                if not df_duree.empty:
                    duree_moyenne = round(df_duree["duree_h"].mean(), 1)

        df_jours = df[df["registration_date"].notna()].copy()
        nb_jours = df_jours["registration_date"].dt.date.nunique()
        dossiers_jour = round(total_dossiers / max(nb_jours, 1), 1)

        # v3.0 : Complexité moyenne globale
        complexite_moy_globale = 0.0
        if "nb_proprietaires_uniques" in df.columns:
            comp_vals = df["nb_proprietaires_uniques"].dropna()
            if not comp_vals.empty:
                complexite_moy_globale = round(comp_vals.mean(), 1)

        top_serie = (
            df.groupby("operateur")["numero_dossier"]
            .nunique()
            .sort_values(ascending=False)
            .head(1)
        )
        top_agent_name = str(top_serie.index[0]) if not top_serie.empty else "—"
        top_agent_count = int(top_serie.iloc[0]) if not top_serie.empty else 0

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine="openpyxl")
        workbook = writer.book

        ws_resume = workbook.create_sheet("Résumé Global", 0)

        ws_resume.merge_cells("A1:H1")
        ws_resume["A1"] = (
            "RAPPORT DE SUIVI GLOBAL - Archivage Numérisation"
        )
        ws_resume["A1"].font = FONT_TITLE
        ws_resume["A1"].alignment = ALIGN_CENTER
        ws_resume.row_dimensions[1].height = 30

        filter_text = f"Produit le : {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        filter_text += f"CSF : {filters.get('csf', 'Tous')}  |  "
        filter_text += f"Bureaux : {filters.get('service', 'Tous')}  |  "
        filter_text += f"Période : {filters.get('start_date', '—')} au {filters.get('end_date', '—')}"
        ws_resume.merge_cells("A2:H2")
        ws_resume["A2"] = filter_text
        ws_resume["A2"].font = FONT_SUBTITLE
        ws_resume["A2"].alignment = ALIGN_CENTER
        ws_resume.row_dimensions[2].height = 20

        # v3.0 : 7 KPIs + Complexité
        kpi_data = [
            ("TOTAL DOSSIERS", str(total_dossiers)),
            ("PIÈCES INDEXÉES", str(int(pieces_indexees))),
            ("TAUX RESTITUTION", f"{taux_restitution} %"),
            ("ATTENTE NUMÉRISATION", str(attente_num)),
            ("DURÉE MOYENNE SORTIE", f"{duree_moyenne} h"),
            ("TOP AGENT", f"{top_agent_name} ({top_agent_count})"),
            ("COMPLEXITÉ MOY.", f"{complexite_moy_globale} prop./dossier"),
        ]

        for i, (label, value) in enumerate(kpi_data):
            col_start = i + 1
            cell_lbl = ws_resume.cell(row=4, column=col_start)
            cell_lbl.value = label
            cell_lbl.font = FONT_KPI_LABEL
            cell_lbl.alignment = ALIGN_CENTER
            cell_val = ws_resume.cell(row=5, column=col_start)
            cell_val.value = value
            cell_val.font = FONT_KPI_VALUE
            cell_val.alignment = ALIGN_CENTER
            for r in range(4, 6):
                c = ws_resume.cell(row=r, column=col_start)
                c.fill = FILL_CREAM
                c.border = BORDER_THIN

        ws_resume.row_dimensions[4].height = 18
        ws_resume.row_dimensions[5].height = 32

        row = 8
        ws_resume.merge_cells(f"A{row}:H{row}")
        ws_resume.cell(row=row, column=1).value = "SYNTHÈSE PAR BUREAU"
        ws_resume.cell(row=row, column=1).font = FONT_SECTION
        ws_resume.cell(row=row, column=1).alignment = ALIGN_LEFT
        row += 1

        # v3.0 : Ajout colonne Complexité Moyenne
        headers_bureau = [
            "Bureau", "Dossiers Traités", "Pièces Indexées",
            "En Attente Num.", "% Réalisation", "Taux Restitution",
            "Complexité Moy.",
        ]
        for j, h in enumerate(headers_bureau, 1):
            ws_resume.cell(row=row, column=j, value=h)
        _apply_header_style(ws_resume, row, len(headers_bureau))
        row += 1

        for bureau in sorted(df["service_origine"].dropna().unique()):
            df_b = df[df["service_origine"] == bureau]
            nb_dossiers = df_b["numero_dossier"].nunique()
            nb_pieces = df_b[df_b["date_indexation"].notna()].drop_duplicates(subset=["numero_dossier"])["count_item"].sum()
            if pd.isna(nb_pieces):
                nb_pieces = 0
            attente_b = df_b[
                (df_b["is_for_scan"] == True) & (df_b["date_indexation"].isna())
            ]["numero_dossier"].nunique()
            rest_b = df_b[df_b["date_retour"].notna()]["numero_dossier"].nunique()
            taux_rest_b = round((rest_b / nb_dossiers * 100), 1) if nb_dossiers > 0 else 0
            pct_real = round((nb_dossiers / max(objectif, 1)) * 100, 1)

            # v3.0 : Complexité par bureau
            comp_b = 0.0
            if "nb_proprietaires_uniques" in df_b.columns:
                comp_vals = df_b["nb_proprietaires_uniques"].dropna()
                if not comp_vals.empty:
                    comp_b = round(comp_vals.mean(), 1)

            vals = [
                bureau, nb_dossiers, int(nb_pieces), attente_b,
                f"{pct_real} %", f"{taux_rest_b} %", comp_b,
            ]
            for j, v in enumerate(vals, 1):
                cell = ws_resume.cell(row=row, column=j, value=v)
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN
            row += 1

        vals_total = [
            "TOTAL", total_dossiers, int(pieces_indexees), attente_num,
            f"{round((total_dossiers / max(objectif, 1)) * 100, 1)} %",
            f"{taux_restitution} %", complexite_moy_globale,
        ]
        for j, v in enumerate(vals_total, 1):
            cell = ws_resume.cell(row=row, column=j, value=v)
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN
            cell.fill = FILL_LIGHT_GOLD

        row += 3
        ws_resume.merge_cells(f"A{row}:H{row}")
        ws_resume.cell(row=row, column=1).value = (
            "PRODUCTION PAR CENTRE DES SERVICES FISCAUX (CSF)"
        )
        ws_resume.cell(row=row, column=1).font = FONT_SECTION
        ws_resume.cell(row=row, column=1).alignment = ALIGN_LEFT
        row += 1

        # v3.0 : Ajout Complexité
        headers_csf = [
            "CSF", "Bureau", "Dossiers Traités",
            "Pièces Indexées", "Taux Restitution", "Dossiers / Jour", "Complexité Moy.",
        ]
        for j, h in enumerate(headers_csf, 1):
            ws_resume.cell(row=row, column=j, value=h)
        _apply_header_style(ws_resume, row, len(headers_csf))
        row += 1

        for csf in sorted(df["csf_geographique"].dropna().unique()):
            for bureau in sorted(
                df[df["csf_geographique"] == csf]["service_origine"].dropna().unique()
            ):
                df_cb = df[
                    (df["csf_geographique"] == csf) & (df["service_origine"] == bureau)
                ]
                nb_d = df_cb["numero_dossier"].nunique()
                nb_p = df_cb[df_cb["date_indexation"].notna()].drop_duplicates(subset=["numero_dossier"])["count_item"].sum()
                if pd.isna(nb_p):
                    nb_p = 0
                rest = df_cb[df_cb["date_retour"].notna()]["numero_dossier"].nunique()
                taux_r = round((rest / nb_d * 100), 1) if nb_d > 0 else 0
                nb_j = df_cb[df_cb["registration_date"].notna()][
                    "registration_date"
                ].dt.date.nunique()
                dpj = round(nb_d / max(nb_j, 1), 1)

                # v3.0 : Complexité par combo CSF+Bureau
                comp_cb = 0.0
                if "nb_proprietaires_uniques" in df_cb.columns:
                    comp_vals = df_cb["nb_proprietaires_uniques"].dropna()
                    if not comp_vals.empty:
                        comp_cb = round(comp_vals.mean(), 1)

                vals = [csf, bureau, nb_d, int(nb_p), f"{taux_r} %", dpj, comp_cb]
                for j, v in enumerate(vals, 1):
                    cell = ws_resume.cell(row=row, column=j, value=v)
                    cell.font = FONT_NORMAL
                    cell.alignment = ALIGN_CENTER
                    cell.border = BORDER_THIN
                row += 1

        _auto_adjust_columns(ws_resume)

        csf_list = sorted(df["csf_geographique"].dropna().unique())
        for csf in csf_list:
            df_csf = df[df["csf_geographique"] == csf]
            sheet_name = csf[:31]
            ws = workbook.create_sheet(sheet_name)
            _write_aggregation_sheet(ws, df_csf, f"PRODUCTION - {csf.upper()}")

        bureau_list = sorted(df["service_origine"].dropna().unique())
        if len(csf_list) > 1:
            for bureau in bureau_list:
                df_bureau = df[df["service_origine"] == bureau]
                sheet_name = f"All {bureau}"[:31]
                ws = workbook.create_sheet(sheet_name)
                _write_aggregation_sheet(ws, df_bureau, f"PRODUCTION - all {bureau.upper()}")

        for csf in csf_list:
            for bureau in bureau_list:
                df_combo = df[
                    (df["csf_geographique"] == csf)
                    & (df["service_origine"] == bureau)
                ]
                if not df_combo.empty:
                    if len(csf_list) == 1 and len(bureau_list) == 1:
                        continue
                    sheet_name = f"{bureau} {csf}"[:31]
                    ws = workbook.create_sheet(sheet_name)
                    _write_aggregation_sheet(ws, df_combo, f"PRODUCTION - {bureau.upper()} {csf.upper()}")

        ws_detail = workbook.create_sheet("Détail Brut")
        df_detail = df.sort_values("registration_date", ascending=False).head(500)

        ws_detail.merge_cells("A1:K1")
        ws_detail["A1"] = "REGISTRE DÉTAILLÉ DES DOSSIERS EXTRAITS"
        ws_detail["A1"].font = FONT_TITLE
        ws_detail["A1"].alignment = ALIGN_CENTER
        ws_detail.row_dimensions[1].height = 28

        ws_detail.merge_cells("A2:K2")
        ws_detail["A2"] = f"500 derniers enregistrements | Export KAGU"
        ws_detail["A2"].font = FONT_SUBTITLE
        ws_detail["A2"].alignment = ALIGN_CENTER

        # v3.0 : Ajout colonnes Vrai Type et Complexité
        detail_headers = [
            "N° Dossier", "CSF", "Bureau", "Opérateur", "Vrai Type",
            "Date Enregistrement", "Date Retour", "Date Indexation",
            "Statut", "Pièces Indexées", "Complexité", "Anomalie Chrono",
        ]
        for j, h in enumerate(detail_headers, 1):
            ws_detail.cell(row=4, column=j, value=h)
        _apply_header_style(ws_detail, 4, len(detail_headers))

        for idx, row_data in df_detail.iterrows():
            r = ws_detail.max_row + 1
            if pd.notna(row_data.get("date_indexation")):
                statut = "Indexé"
            elif pd.notna(row_data.get("registration_date")):
                statut = "En cours"
            else:
                statut = "—"

            anomalie = "Oui" if row_data.get("anomalie_chronologie") == True else "Non"
            vrai_type = row_data.get("vrai_type", "—")
            complexite = int(row_data.get("nb_proprietaires_uniques", 0)) if pd.notna(row_data.get("nb_proprietaires_uniques")) else 0

            vals = [
                row_data.get("numero_dossier", "—"),
                row_data.get("csf_geographique", "—"),
                row_data.get("service_origine", "—"),
                row_data.get("operateur", "—"),
                vrai_type,
                row_data["registration_date"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("registration_date")) else "—",
                row_data["date_retour"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("date_retour")) else "—",
                row_data["date_indexation"].strftime("%d/%m/%Y %H:%M")
                if pd.notna(row_data.get("date_indexation")) else "—",
                statut,
                int(row_data.get("count_item", 0)) if pd.notna(row_data.get("count_item")) else 0,
                complexite,
                anomalie,
            ]
            for j, v in enumerate(vals, 1):
                cell = ws_detail.cell(row=r, column=j, value=v)
                cell.font = FONT_NORMAL
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN
                # v3.0 : Coloration conditionnelle Vrai Type
                if j == 5:  # Colonne Vrai Type
                    if v == "Entrée":
                        cell.fill = FILL_GREEN_LIGHT
                        cell.font = Font(name="Calibri", size=11, bold=True, color=GREEN_TEXT.lstrip("#"))
                    elif v == "Sortie":
                        cell.fill = FILL_RED_LIGHT
                        cell.font = Font(name="Calibri", size=11, bold=True, color=RED_TEXT.lstrip("#"))

        _auto_adjust_columns(ws_detail)

        writer.close()
        output.seek(0)
        return output.read()

    finally:
        con.close()


# =============================================================================
# DASHBOARD STATE
# =============================================================================
class DashboardState(rx.State):
    start_date: str = ""
    end_date: str = ""
    selected_service: str = "Tous"
    selected_csf: str = "Tous"
    selected_type: str = "Tous"
    flux_period: str = "Mois"
    is_loading: bool = False
    objectif_attendu: int = 100
    db_error: str = ""
    type_options: list[str] = ["Tous", "Entrée", "Sortie"]  # v3.0 : vrai_type

    kpi_objectif: int = 100
    kpi_taux_indexation: float = 0.0
    kpi_taux_num: float = 0.0
    kpi_total_dossiers: int = 0
    kpi_attente_num: int = 0
    kpi_pieces_indexees: int = 0
    kpi_taux_restitution: float = 0.0
    kpi_duree_moyenne: float = 0.0
    kpi_dossiers_jour: float = 0.0
    kpi_top_archiviste: str = "—"
    kpi_top_total: int = 0
    kpi_bar_indexation: float = 0.0
    kpi_bar_num: float = 0.0
    kpi_ratio_indexation: float = 0.0
    # v3.0 : Nouveaux KPIs complexité
    kpi_complexite_moyenne: float = 0.0
    kpi_complexite_conservation: float = 0.0
    kpi_complexite_cadastre: float = 0.0
    kpi_complexite_domaines: float = 0.0

    flux_data: list[dict] = []
    flux_svg_html: str = ""
    volumes_bureau: list[dict] = []
    donut_html: str = ""
    productivite: list[dict] = []
    productivite_count: int = 0
    derniers_enregistrements: list[dict] = []
    # v3.0 : Données de complexité par bureau
    complexite_bureau: list[dict] = []
    # v3.0 : Stats entrée/sortie
    stats_es: dict = {"entrees": 0, "sorties": 0, "sans_sortie": 0}

    # =================================================================
    # v2.8.7 — Graphique flux SVG
    # =================================================================
    def _build_flux_svg(self, data_list: list[dict]) -> str:
        if not data_list:
            return '<svg width="100%" height="100%" viewBox="0 0 600 300" xmlns="http://www.w3.org/2000/svg"><text x="50%" y="50%" text-anchor="middle" fill="#AAA" font-size="14" font-family="sans-serif">Aucune donnée</text></svg>'

        svg_w, svg_h = 600, 240
        m_left, m_right = 50, 40
        m_top, m_bottom = 30, 80

        chart_w = svg_w - m_left - m_right
        chart_h = svg_h - m_top - m_bottom

        max_val = max(d["value"] for d in data_list) * 1.15 or 100
        n = len(data_list)

        points = []
        for i, d in enumerate(data_list):
            if n > 1:
                x = m_left + (i * chart_w / (n - 1))
            else:
                x = m_left + chart_w / 2
            y = m_top + chart_h - (d["value"] / max_val * chart_h)
            points.append((x, y))

        path_d = f"M {points[0][0]:.1f} {points[0][1]:.1f}"
        for x, y in points[1:]:
            path_d += f" L {x:.1f} {y:.1f}"

        fill_d = path_d + f" L {points[-1][0]:.1f} {m_top + chart_h:.1f} L {points[0][0]:.1f} {m_top + chart_h:.1f} Z"

        grid_lines = ""
        y_labels = ""
        for i in range(5):
            ratio = i / 4.0
            y_line = m_top + chart_h * (1 - ratio)
            val_label = round(max_val * ratio)
            grid_lines += f'<line x1="{m_left}" y1="{y_line:.1f}" x2="{m_left + chart_w}" y2="{y_line:.1f}" stroke="#F0F0F0" stroke-width="1"/>'
            y_labels += f'<text x="{m_left - 8}" y="{y_line + 3:.1f}" text-anchor="end" font-size="9" fill="#AAA">{val_label}</text>'

        x_labels = ""
        for i, d in enumerate(data_list):
            x_pos = points[i][0]
            label_y = svg_h - m_bottom + 10
            x_labels += f'''<text x="{x_pos:.1f}" y="{label_y}" font-size="10" fill="#888" 
                            transform="rotate(-45, {x_pos:.1f}, {label_y})" 
                            text-anchor="end">{d["date"]}</text>'''

        circles = ""
        for x, y in points:
            circles += f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{GOLD}" stroke="{WHITE}" stroke-width="2"/>'

        return f'''<svg width="100%" height="100%" viewBox="0 0 {svg_w} {svg_h}" xmlns="http://www.w3.org/2000/svg">
            <defs>
                <linearGradient id="areaGrad" x1="0" y1="0" x2="0" y2="1">
                    <stop offset="0%" stop-color="{GOLD}" stop-opacity="0.2"/>
                    <stop offset="100%" stop-color="{GOLD}" stop-opacity="0.0"/>
                </linearGradient>
            </defs>
            {grid_lines}{y_labels}
            <path d="{fill_d}" fill="url(#areaGrad)" stroke="none"/>
            <path d="{path_d}" fill="none" stroke="{GOLD}" stroke-width="2.5" stroke-linecap="round"/>
            {circles}{x_labels}
        </svg>'''

    # =================================================================
    # v2.8.4 — Donut HTML
    # =================================================================
    def _build_donut_html(self, vol_data: list[dict]) -> str:
        if not vol_data:
            return '<div style="display:flex;align-items:center;justify-content:center;height:100%;color:#AAA;font-family:sans-serif;">Aucune donnée</div>'

        total = sum(d["volume"] for d in vol_data)
        cx, cy, r = 140, 140, 88
        circumference = 2 * math.pi * r
        circles_html = ""
        offset = 0

        for d in vol_data:
            dash = (d["volume"] / total) * circumference
            gap = circumference - dash
            circles_html += f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{d["color"]}" stroke-width="32" stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})" stroke-linecap="butt"/>'
            offset += dash

        legend_items = ""
        for d in vol_data:
            legend_items += f'''
            <div style="display:flex;align-items:flex-start;gap:8px;margin-bottom:8px;font-family:sans-serif;">
                <div style="width:10px;height:10px;border-radius:2px;background:{d["color"]};margin-top:3px;flex-shrink:0;"></div>
                <div style="flex:1;min-width:0;">
                    <div style="font-size:12px;color:#333;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{d["bureau"]}</div>
                    <div style="font-size:11px;color:#8A8A8A;">
                        <span style="font-weight:600;color:#1E1E1E;">{d["volume"]}</span>
                        <span style="margin-left:4px;">{d["pct"]}%</span>
                    </div>
                </div>
            </div>'''

        return f'''
        <div style="display:flex;align-items:center;width:100%;height:100%;padding:28px;box-sizing:border-box;gap:28px;">
            <div style="position:relative;width:300px;height:300px;flex-shrink:0;">
                <svg width="300" height="300" viewBox="0 0 280 280">
                    {circles_html}
                </svg>
                <div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;">
                    <div style="font-size:24px;font-weight:bold;color:#1E1E1E;font-family:sans-serif;">{total}</div>
                    <div style="font-size:10px;color:#8A8A8A;font-family:sans-serif;">Total</div>
                </div>
            </div>
            <div style="flex:1;min-width:120px;max-width:220px;">
                {legend_items}
            </div>
        </div>'''

    async def load_data(self):
        self.is_loading = True
        self.db_error = ""
        try:
            await asyncio.to_thread(self._fetch_all_data)
        except Exception as e:
            self.db_error = f"Erreur base de données : {str(e)}"
            print(self.db_error)
        finally:
            self.is_loading = False

    def _fetch_all_data(self):
        if not os.path.exists(DUCKDB_PATH):
            raise FileNotFoundError(f"Base DuckDB introuvable : {DUCKDB_PATH}")
        con = duckdb.connect(DUCKDB_PATH, read_only=True)
        try:
            filters = {"start_date": self.start_date, "end_date": self.end_date, "service": self.selected_service, "csf": self.selected_csf, "type": self.selected_type}
            where_sql, params = _build_where(filters)

            # v3.0 : Vérifier que fact_suivi_global a les colonnes nécessaires
            # Si pas de colonnes vrai_type / nb_proprietaires_uniques, on les simule
            cols_info = con.execute("PRAGMA table_info(fact_suivi_global)").fetchall()
            col_names = [c[1] for c in cols_info]

            has_vrai_type = "vrai_type" in col_names
            has_complexite = "nb_proprietaires_uniques" in col_names
            has_record_type_uuid = "record_type_uuid" in col_names
            has_output_id = "output_id" in col_names
            has_duree_num = "duree_numerisation_h" in col_names

            # Adapter les requêtes selon la structure disponible
            if not has_vrai_type:
                # Fallback : utiliser motif_enregistrement comme proxy
                # mais logiquement on devrait reconstruire la vue
                print("AVERTISSEMENT: colonne vrai_type absente, utilisation de motif_enregistrement")

            # --- KPI Total Dossiers ---
            row = con.execute(f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE {where_sql}", params).fetchone()
            self.kpi_total_dossiers = _safe_int(row[0])

            # --- KPI Attente Numérisation ---
            row = con.execute(f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE is_for_scan = TRUE AND date_indexation IS NULL AND {where_sql}", params).fetchone()
            self.kpi_attente_num = _safe_int(row[0])

            # --- KPI Pièces Indexées ---
            # v3.1 : count_item est un attribut par DOSSIER (general_index), pas par
            # mouvement. Il faut le dédupliquer par numero_dossier avant SUM, sinon
            # il est compté une fois par mouvement (Entrée/Sortie/Retour).
            row = con.execute(
                f"""
                SELECT COALESCE(SUM(count_item), 0) FROM (
                    SELECT numero_dossier, MAX(count_item) AS count_item
                    FROM fact_suivi_global
                    WHERE date_indexation IS NOT NULL AND {where_sql}
                    GROUP BY numero_dossier
                )
                """,
                params
            ).fetchone()
            self.kpi_pieces_indexees = _safe_int(row[0])

            # --- KPI Taux Restitution ---
            row = con.execute(f"SELECT COALESCE(COUNT(DISTINCT CASE WHEN date_retour IS NOT NULL THEN numero_dossier END) * 100.0 / NULLIF(COUNT(DISTINCT numero_dossier), 0), 0) FROM fact_suivi_global WHERE {where_sql}", params).fetchone()
            self.kpi_taux_restitution = round(_safe_float(row[0]), 1)

            # --- KPI Durée Moyenne Numérisation ---
            # v3.1 : durée = Entrée de retour liée (output_id) - Sortie envoi scan,
            # calculée dans l'ETL via duree_numerisation_h. Fallback sur
            # date_retour - date_indexation si l'ETL n'a pas encore été rejoué.
            if has_duree_num:
                row = con.execute(
                    f"""
                    SELECT COALESCE(AVG(TRY_CAST(duree_numerisation_h AS DOUBLE)), 0)
                    FROM fact_suivi_global
                    WHERE duree_numerisation_h IS NOT NULL AND {where_sql}
                    """,
                    params
                ).fetchone()
            else:
                # --- KPI Durée Moyenne Numérisation ---
                # v3.2 : utilise directement duree_numerisation_h calculée par l'ETL
                row = con.execute(
                    f"""
                    SELECT COALESCE(AVG(duree_numerisation_h), 0)
                    FROM fact_suivi_global
                    WHERE duree_numerisation_h IS NOT NULL AND {where_sql}
                    """,
                    params
                ).fetchone()

            self.kpi_duree_moyenne = round(_safe_float(row[0]), 1)

            # --- KPI Dossiers / Jour ---
            row = con.execute(f"SELECT COALESCE(COUNT(DISTINCT numero_dossier) * 1.0 / NULLIF(COUNT(DISTINCT registration_date::DATE), 0), 0) FROM fact_suivi_global WHERE {where_sql}", params).fetchone()
            self.kpi_dossiers_jour = round(_safe_float(row[0]), 1)

            # --- KPI Taux Indexation ---
            row = con.execute(f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE date_indexation IS NOT NULL AND {where_sql}", params).fetchone()
            indexes = _safe_int(row[0])
            self.kpi_taux_indexation = round((indexes / max(self.objectif_attendu, 1)) * 100, 1)

            # --- KPI Taux Numérisation ---
            row = con.execute(f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE is_for_scan = TRUE AND {where_sql}", params).fetchone()
            numerises = _safe_int(row[0])
            self.kpi_taux_num = round((numerises / max(self.objectif_attendu, 1)) * 100, 1)

            self.kpi_bar_indexation = min(self.kpi_taux_indexation, 100.0)
            self.kpi_bar_num = min(self.kpi_taux_num, 100.0)
            self.kpi_ratio_indexation = round(self.kpi_taux_indexation / 100.0, 1)

            # --- KPI Top Archiviste ---
            row = con.execute(f"SELECT operateur, COUNT(DISTINCT numero_dossier) as total FROM fact_suivi_global WHERE {where_sql} GROUP BY operateur ORDER BY total DESC LIMIT 1", params).fetchone()
            if row and row[0]:
                self.kpi_top_archiviste = str(row[0])
                self.kpi_top_total = _safe_int(row[1])
            else:
                self.kpi_top_archiviste = "—"
                self.kpi_top_total = 0

            # v3.0 --- KPI Complexité Moyenne Globale ---
            if has_complexite:
                row = con.execute(
                    f"""
                    SELECT COALESCE(AVG(nb_proprietaires_uniques), 0)
                    FROM (
                        SELECT numero_dossier, MAX(nb_proprietaires_uniques) as nb_proprietaires_uniques
                        FROM fact_suivi_global
                        WHERE {where_sql}
                        GROUP BY numero_dossier
                    )
                    """, params
                ).fetchone()
                self.kpi_complexite_moyenne = round(_safe_float(row[0]), 1)

                # Complexité par bureau
                rows = con.execute(
                    f"""
                    SELECT service_origine, COALESCE(AVG(nb_prop), 0) as comp_moy
                    FROM (
                        SELECT service_origine, numero_dossier, MAX(nb_proprietaires_uniques) as nb_prop
                        FROM fact_suivi_global
                        WHERE service_origine IS NOT NULL AND {where_sql}
                        GROUP BY service_origine, numero_dossier
                    )
                    GROUP BY service_origine
                    ORDER BY comp_moy DESC
                    """, params
                ).fetchall()
                # self.complexite_bureau = [
                #     {"bureau": str(r[0]), "complexite": round(_safe_float(r[1]), 1)} 
                #     for r in rows
                # ]
                self.complexite_bureau = [
                    {
                        "bureau": str(r[0]),
                        "complexite": round(_safe_float(r[1]), 1),
                        "bar_width_str": f"{min(round(_safe_float(r[1]) * 20), 100)}%",
                        "categorie": "Élevée" if round(_safe_float(r[1]), 1) >= 5 else ("Moyenne" if round(_safe_float(r[1]), 1) >= 2 else "Faible"),
                        "is_conservation": str(r[0]) == "Conservation",
                        "is_elevee": "Élevée" if round(_safe_float(r[1]), 1) >= 5 else ("Moyenne" if round(_safe_float(r[1]), 1) >= 2 else "Faible") == "Élevée",
                        "is_moyenne": "Élevée" if round(_safe_float(r[1]), 1) >= 5 else ("Moyenne" if round(_safe_float(r[1]), 1) >= 2 else "Faible") == "Moyenne",
                    }
                    for r in rows
                ]

                # Complexité par bureau (KPIs individuels)
                for r in rows:
                    bureau = str(r[0])
                    comp = round(_safe_float(r[1]), 1)
                    if bureau == "Conservation":
                        self.kpi_complexite_conservation = comp
                    elif bureau == "Cadastre":
                        self.kpi_complexite_cadastre = comp
                    elif bureau == "Domaines":
                        self.kpi_complexite_domaines = comp
            else:
                self.kpi_complexite_moyenne = 0.0
                self.complexite_bureau = []

            # v3.0 --- Stats Entrée/Sortie ---
            if has_record_type_uuid:
                row_e = con.execute(
                    f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE record_type_uuid = '{UUID_ENTREE}' AND {where_sql}", 
                    params
                ).fetchone()
                row_s = con.execute(
                    f"SELECT COUNT(DISTINCT numero_dossier) FROM fact_suivi_global WHERE record_type_uuid = '{UUID_SORTIE}' AND {where_sql}", 
                    params
                ).fetchone()
                self.stats_es = {
                    "entrees": _safe_int(row_e[0]),
                    "sorties": _safe_int(row_s[0]),
                }
            else:
                # Lève une erreur claire plutôt que de bidouiller avec des chaînes de caractères
                self.db_error = "La colonne record_type_uuid est manquante dans la table fact_suivi_global."
                self.stats_es = {"entrees": 0, "sorties": 0}

            # --- Flux d'activité ---
            if self.flux_period == "Mois":
                period_expr = "STRFTIME('%Y-%m', registration_date::TIMESTAMP)"
            else:
                period_expr = "STRFTIME('%Y', registration_date::TIMESTAMP) || '-Q' || QUARTER(registration_date::TIMESTAMP)"

            flux_sql = (
                f"SELECT {period_expr} as periode, COUNT(DISTINCT numero_dossier) as value "
                f"FROM fact_suivi_global WHERE registration_date IS NOT NULL AND {where_sql} "
                f"GROUP BY {period_expr} ORDER BY {period_expr}"
            )
            rows = con.execute(flux_sql, params).fetchall()
            self.flux_data = [{"date": str(r[0]), "value": _safe_int(r[1])} for r in rows]
            self.flux_svg_html = self._build_flux_svg(self.flux_data)

            # --- Volume par bureau (Donut) ---
            rows = con.execute(f"SELECT service_origine as bureau, COUNT(DISTINCT numero_dossier) as volume FROM fact_suivi_global WHERE service_origine IS NOT NULL AND {where_sql} GROUP BY service_origine ORDER BY volume DESC", params).fetchall()
            total_vol = sum(r[1] for r in rows) if rows else 1
            colors = {"Cadastre": GOLD, "Conservation": BROWN, "Domaines": "#A0826D"}
            self.volumes_bureau = [{"bureau": str(r[0]), "volume": _safe_int(r[1]), "pct": round(_safe_int(r[1]) / total_vol * 100, 1), "color": colors.get(str(r[0]), "#999")} for r in rows]
            self.donut_html = self._build_donut_html(self.volumes_bureau)

            # --- Productivité ---
            rows = con.execute(f"SELECT operateur, service_origine as bureau, COUNT(DISTINCT numero_dossier) as dossiers, COALESCE(COUNT(DISTINCT numero_dossier) * 1.0 / NULLIF(COUNT(DISTINCT registration_date::DATE), 0), 0) as moyenne FROM fact_suivi_global WHERE operateur IS NOT NULL AND {where_sql} GROUP BY operateur, service_origine ORDER BY dossiers DESC LIMIT 11", params).fetchall()
            max_dossiers = max(_safe_int(r[2]) for r in rows) if rows else 1
            self.productivite = [{"rang": i + 1, "agent": str(r[0]), "csf": self.selected_csf if self.selected_csf != "Tous" else "Dakar Plateau", "bureau": str(r[1]), "dossiers": _safe_int(r[2]), "moyenne": round(_safe_float(r[3]), 1), "pct": round((_safe_int(r[2]) / max_dossiers) * 100, 1)} for i, r in enumerate(rows)]
            self.productivite_count = len(self.productivite)

            # --- Derniers enregistrements ---
            # v3.0 : Utiliser vrai_type si disponible
            type_col = "vrai_type" if has_vrai_type else "motif_enregistrement"
            rows = con.execute(f"SELECT numero_dossier as code, COALESCE(NULLIF(folder_label, ''), motif_enregistrement, '—') as desc, COALESCE({type_col}, '—') as type, registration_date FROM fact_suivi_global WHERE {where_sql} ORDER BY registration_date DESC LIMIT 5", params).fetchall()
            self.derniers_enregistrements = [{"code": str(r[0]), "desc": str(r[1]), "type_badge": "Entrée" if "entrée" in str(r[2]).lower() or str(r[2]) == "Entrée" else "Sortie"} for r in rows]
        finally:
            con.close()

    # =================================================================
    # v3.0 — VALIDATION DATES
    # =================================================================
    async def set_start_date(self, value: str):
        self.start_date = value
        valid, msg = _validate_date_range(self.start_date, self.end_date)
        if not valid:
            self.db_error = msg
            self.is_loading = False
            return
        self.db_error = ""
        await self.load_data()

    async def set_end_date(self, value: str):
        self.end_date = value
        valid, msg = _validate_date_range(self.start_date, self.end_date)
        if not valid:
            self.db_error = msg
            self.is_loading = False
            return
        self.db_error = ""
        await self.load_data()

    async def set_service(self, value: str):
        self.selected_service = value
        await self.load_data()

    async def set_csf(self, value: str):
        self.selected_csf = value
        await self.load_data()

    async def set_type(self, value: str):
        self.selected_type = value
        await self.load_data()

    async def reset_service(self):
        self.selected_service = "Tous"
        await self.load_data()

    async def reset_csf(self):
        self.selected_csf = "Tous"
        await self.load_data()

    async def reset_type(self):
        self.selected_type = "Tous"
        await self.load_data()

    async def select_current_week(self):
        today = datetime.today()
        start = today - timedelta(days=today.weekday())
        self.start_date = start.strftime("%Y-%m-%d")
        self.end_date = today.strftime("%Y-%m-%d")
        self.db_error = ""
        await self.load_data()

    async def select_last_recorded_week(self):
        if not os.path.exists(DUCKDB_PATH):
            self.db_error = "Base de données inaccessible"
            return
        con = duckdb.connect(DUCKDB_PATH, read_only=True)
        try:
            max_date_row = con.execute(
                "SELECT MAX(registration_date::DATE) FROM fact_suivi_global WHERE registration_date IS NOT NULL"
            ).fetchone()
            if max_date_row and max_date_row[0]:
                last_recorded = max_date_row[0]
                start = last_recorded - timedelta(days=last_recorded.weekday())
                end = start + timedelta(days=6)
                self.start_date = start.strftime("%Y-%m-%d")
                self.end_date = end.strftime("%Y-%m-%d")
            else:
                today = datetime.today()
                last_week = today - timedelta(days=7)
                start = last_week - timedelta(days=last_week.weekday())
                end = start + timedelta(days=6)
                self.start_date = start.strftime("%Y-%m-%d")
                self.end_date = end.strftime("%Y-%m-%d")
        finally:
            con.close()
        self.db_error = ""
        await self.load_data()

    async def reset_date_filters(self):
        self.start_date = ""
        self.end_date = ""
        self.db_error = ""
        await self.load_data()

    async def set_flux_mois(self):
        self.flux_period = "Mois"
        await self.load_data()

    async def set_flux_trimestre(self):
        self.flux_period = "Trimestre"
        await self.load_data()

    async def set_objectif_attendu(self, value: str):
        try:
            self.objectif_attendu = int(value)
            self.kpi_objectif = int(value)
        except ValueError:
            pass
        await self.load_data()

    async def on_load(self):
        await self.select_current_week()

    async def export_excel(self):
        """Génère et télécharge le rapport Excel multi-feuilles."""
        self.is_loading = True
        self.db_error = ""
        try:
            filters = {
                "start_date": self.start_date,
                "end_date": self.end_date,
                "service": self.selected_service,
                "csf": self.selected_csf,
                "type": self.selected_type,
            }
            data = await asyncio.to_thread(
                generate_excel_report,
                DUCKDB_PATH,
                filters,
                self.objectif_attendu,
            )
            filename = f"suivi_dgid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return rx.download(data=data, filename=filename)
        except Exception as e:
            self.db_error = f"Erreur export Excel : {str(e)}"
            print(self.db_error)
        finally:
            self.is_loading = False


# =============================================================================
# COMPONENTS (Vue)
# =============================================================================

def bureau_badge(bureau: str) -> rx.Component:
    bg = rx.cond(bureau == "Cadastre", CADASTRE_BG, rx.cond(bureau == "Conservation", CONSERVATION_BG, DOMAINES_BG))
    color = rx.cond(bureau == "Cadastre", CADASTRE_TEXT, rx.cond(bureau == "Conservation", CONSERVATION_TEXT, DOMAINES_TEXT))
    return rx.badge(bureau, bg=bg, color=color, border_radius="6px", padding_x="10px", padding_y="2px", font_size="11px", font_weight="500")


def type_badge(type_badge_str: str) -> rx.Component:
    bg = rx.cond(type_badge_str == "Entrée", GREEN_LIGHT, RED_LIGHT)
    color = rx.cond(type_badge_str == "Entrée", GREEN_TEXT, RED_TEXT)
    return rx.badge(type_badge_str, bg=bg, color=color, border_radius="20px", padding_x="12px", padding_y="3px", font_size="11px", font_weight="500")


def header() -> rx.Component:
    return rx.hstack(
        rx.hstack(
            rx.box(rx.icon("landmark", size=22, color=GOLD), bg=ICON_BG, padding="10px", border_radius="10px", display="flex", align_items="center", justify_content="center"),
            rx.vstack(rx.heading("Dashboard de Suivi", size="5", color=TEXT_MAIN, font_weight="bold", line_height="1.2"), rx.text("Numérisation & Indexation — sites pilotes CSF", color=TEXT_MUTED, font_size="12px", line_height="1.2"), spacing="1", align_items="start"),
            spacing="3", align_items="center",
        ),
        rx.spacer(),
        rx.button(
            rx.hstack(
                rx.icon("download", size=14),
                rx.text("Exporter en Excel", font_size="13px", font_weight="500"),
                spacing="2", align_items="center"
            ),
            bg=WHITE, color=BROWN, border=f"1px solid {BROWN}", border_radius="8px",
            padding="8px 18px", height="40px", cursor="pointer",
            _hover={"bg": BROWN, "color": WHITE, "border_color": BROWN, "transition": "all 0.2s ease"},
            on_click=DashboardState.export_excel
        ),
        width="100%", padding="16px 32px", bg=WHITE, border_bottom=f"1px solid {BORDER_HDR}", align_items="center",
    )


def filter_bar() -> rx.Component:
    label_style = {"color": TEXT_MUTED, "font_size": "10px", "font_weight": "600", "letter_spacing": "0.5px", "text_transform": "uppercase"}
    input_style = {"border_radius": "8px", "border": f"1px solid {BORDER_CARD}", "padding": "8px 12px", "font_size": "13px", "color": TEXT_MAIN, "width": "140px", "height": "38px"}
    btn_ghost = {"variant": "ghost", "size": "2", "color": BROWN, "font_size": "12px", "border_radius": "20px", "border": f"1px solid {BORDER_CARD}", "padding": "8px 16px", "height": "38px", "cursor": "pointer", "_hover": {"bg": BROWN, "color": WHITE, "transition": "all 0.2s ease"}}
    btn_muted = {"variant": "ghost", "size": "2", "color": TEXT_MUTED, "font_size": "12px", "border_radius": "20px", "border": f"1px solid {BORDER_CARD}", "padding": "8px 16px", "height": "38px", "cursor": "pointer", "_hover": {"color": BROWN, "border_color": BROWN, "transition": "all 0.2s ease"}}
    #select_style = {"border_radius": "8px", "border": f"1px solid {BORDER_CARD}", "font_size": "13px", "width": "130px", "height": "38px", "bg": WHITE, "color": TEXT_MAIN}
    select_style = {
        "background_color": "white",
        "color": "#1A1A1A",
        "border": "1px solid #CBD5E0",
        "border_radius": "0.375rem",
        "box_shadow": "0 1px 2px 0 rgba(0, 0, 0, 0.05)",
        "font_weight": "500",
        "_hover": {
            "border_color": "#718096"
        }
    }


    return rx.box(
        rx.hstack(
            rx.hstack(
                rx.vstack(rx.text("DU", **label_style), rx.input(type="date", value=DashboardState.start_date, on_change=DashboardState.set_start_date, **input_style), spacing="1", align_items="start"),
                rx.vstack(rx.text("AU", **label_style), rx.input(type="date", value=DashboardState.end_date, on_change=DashboardState.set_end_date, **input_style), spacing="1", align_items="start"),
                rx.button("Semaine actuelle", on_click=DashboardState.select_current_week, **btn_ghost),
                rx.button("Dernière semaine", on_click=DashboardState.select_last_recorded_week, **btn_ghost),
                rx.button("Réinitialiser", on_click=DashboardState.reset_date_filters, **btn_muted),
                spacing="4",
                align_items="end",
            ),
            rx.spacer(),
            rx.hstack(
                rx.vstack(rx.text("BUREAU", **label_style), rx.select(["Tous", "Cadastre", "Conservation", "Domaines"], value=DashboardState.selected_service, on_change=DashboardState.set_service,variant="surface", color_scheme="gray", width="130px", height="38px", style={"font_size": "13px"}), spacing="1", align_items="start"), #**select_style), 
                rx.vstack(rx.text("CSF", **label_style), rx.select(["Tous", "Dakar Plateau", "Ngor-Almadies", "Mbour"], value=DashboardState.selected_csf, on_change=DashboardState.set_csf, variant="surface",  color_scheme="gray", width="130px", height="38px", style={"font_size": "13px"}),spacing="1", align_items="start"), #**select_style), spacing="1", align_items="start"),
                rx.vstack(rx.text("TYPE", **label_style), rx.select(DashboardState.type_options, value=DashboardState.selected_type, on_change=DashboardState.set_type, **select_style), spacing="1", align_items="start"),
                spacing="5",
                align_items="end",
            ),
            width="100%",
            align_items="end",
        ),
        padding="14px 32px",
        bg=WHITE,
        border_bottom=f"1px solid {BORDER_HDR}",
        width="100%",
    )


def objectives_section() -> rx.Component:
    return rx.vstack(
        rx.hstack(
            rx.hstack(rx.icon("target", size=14, color=BROWN), rx.text("PILOTAGE SUPERVISEUR", color=BROWN, font_size="10px", font_weight="700", letter_spacing="1px"), spacing="2", align_items="center"),
            rx.spacer(),
            rx.vstack(
                rx.text("DOSSIERS ATTENDUS", color=TEXT_MUTED, font_size="10px", font_weight="600", letter_spacing="0.5px"), 
                rx.input(
                    value=DashboardState.objectif_attendu.to_string(), 
                    on_change=DashboardState.set_objectif_attendu, 
                    width="100px", height="32px", 
                    font_size="14px", font_weight="600", text_align="center", 
                    color=TEXT_MAIN, bg=WHITE, border_radius="8px", border=f"1px solid {BORDER_CARD}"
                ), 
                spacing="1", align_items="end"
            ),
            width="100%", align_items="start", margin_bottom="8px",
        ),
        rx.heading("Objectifs de la période", size="4", color=TEXT_MAIN, font_weight="bold", margin_bottom="16px"),
        rx.grid(
            rx.box(rx.vstack(rx.text("Objectif fixé", color=TEXT_MUTED, font_size="12px", font_weight="500"), rx.hstack(rx.text(DashboardState.kpi_objectif.to_string(), font_size="36px", font_weight="bold", color=TEXT_MAIN, line_height="1"), rx.text("dossiers", font_size="14px", color=TEXT_MUTED, margin_top="10px"), spacing="2", align_items="end", margin_top="12px"), rx.text("Objectif de la période", color=TEXT_MUTED, font_size="10px", margin_top="8px"), spacing="1", align_items="center", justify_content="center", height="100%"), padding="24px", bg=CREAM_CARD, border_radius="12px", border=f"1px solid #F5E6C8", width="100%", height="100%"),
            rx.box(rx.vstack(rx.hstack(rx.text("Taux d'indexation", color=TEXT_MUTED, font_size="12px", font_weight="500"), rx.spacer(), rx.box(rx.icon("layers", size=16, color=GOLD), bg=ICON_BG, padding="6px", border_radius="8px"), width="100%", align_items="center"), rx.hstack(rx.text(f"{DashboardState.kpi_taux_indexation}", font_size="28px", font_weight="bold", color=TEXT_MAIN, line_height="1"), rx.text("%", font_size="16px", color=TEXT_MUTED, margin_top="4px"), spacing="1", align_items="end"), rx.box(rx.box(width=f"{DashboardState.kpi_bar_indexation}%", height="6px", bg=GOLD, border_radius="3px"), width="100%", bg="#F0E6D0", border_radius="3px", height="6px", margin_top="8px"), rx.hstack(rx.text("Indexés / objectif de la période", color=TEXT_MUTED, font_size="10px"), rx.spacer(), rx.text(f"×{DashboardState.kpi_ratio_indexation}", color=GOLD, font_size="10px", font_weight="bold"), width="100%", margin_top="4px"), spacing="2", align_items="start", width="100%"), padding="20px", bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%", height="100%"),
            rx.box(rx.vstack(rx.hstack(rx.text("Taux de numérisation", color=TEXT_MUTED, font_size="12px", font_weight="500"), rx.spacer(), rx.box(rx.icon("scan", size=16, color=GOLD), bg=ICON_BG, padding="6px", border_radius="8px"), width="100%", align_items="center"), rx.hstack(rx.text(f"{DashboardState.kpi_taux_num}", font_size="28px", font_weight="bold", color=TEXT_MAIN, line_height="1"), rx.text("%", font_size="16px", color=TEXT_MUTED, margin_top="4px"), spacing="1", align_items="end"), rx.box(rx.box(width=f"{DashboardState.kpi_bar_num}%", height="6px", bg=GOLD, border_radius="3px"), width="100%", bg="#F0E6D0", border_radius="3px", height="6px", margin_top="8px"), rx.text("Numérisés / objectif de la période", color=TEXT_MUTED, font_size="10px", margin_top="4px"), spacing="2", align_items="start", width="100%"), padding="20px", bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%", height="100%"),
            columns="3", spacing="4", width="100%",
        ),
        width="100%", spacing="0",
    )


def metrics_section() -> rx.Component:
    def std_card(title, value, subtitle, icon_name):
        return rx.box(rx.vstack(rx.hstack(rx.text(title, color=TEXT_MUTED, font_size="12px", font_weight="500"), rx.spacer(), rx.box(rx.icon(icon_name, size=16, color=GOLD), bg=ICON_BG, padding="6px", border_radius="8px"), width="100%", align_items="center"), rx.text(value, font_size="24px", font_weight="bold", color=TEXT_MAIN, line_height="1", margin_top="8px"), rx.text(subtitle, color=TEXT_MUTED, font_size="11px", margin_top="4px"), spacing="1", align_items="start", width="100%"), padding="18px", bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%", height="100%")
    def dark_card(title, value, subtitle, icon_name):
        return rx.box(rx.vstack(rx.hstack(rx.text(title, color="rgba(255,255,255,0.7)", font_size="12px", font_weight="500"), rx.spacer(), rx.box(rx.icon(icon_name, size=16, color=GOLD), bg="rgba(255,255,255,0.1)", padding="6px", border_radius="8px"), width="100%", align_items="center"), rx.text(value, font_size="24px", font_weight="bold", color=WHITE, line_height="1", margin_top="8px"), rx.text(subtitle, color="rgba(255,255,255,0.7)", font_size="11px", margin_top="4px"), spacing="1", align_items="start", width="100%"), padding="18px", bg=BROWN_DARK, border_radius="12px", width="100%", height="100%")

    # v3.0 : Ajout carte complexité moyenne
    def complexite_card(title, value, subtitle, icon_name, bureau_name=""):
        # Couleur spécifique pour Conservation (plus complexe)
        bg_color = rx.cond(bureau_name == "Conservation", "#F5E6C8", CREAM_CARD)
        txt_color = rx.cond(bureau_name == "Conservation", BROWN, TEXT_MAIN)
        return rx.box(
            rx.vstack(
                rx.hstack(
                    rx.text(title, color=TEXT_MUTED, font_size="12px", font_weight="500"),
                    rx.spacer(),
                    rx.box(rx.icon(icon_name, size=16, color=GOLD), bg=ICON_BG, padding="6px", border_radius="8px"),
                    width="100%", align_items="center"
                ),
                rx.hstack(
                    rx.text(value, font_size="24px", font_weight="bold", color=txt_color, line_height="1"),
                    rx.text("prop./dossier", font_size="11px", color=TEXT_MUTED, margin_top="4px"),
                    spacing="1", align_items="end"
                ),
                rx.text(subtitle, color=TEXT_MUTED, font_size="11px", margin_top="4px"),
                spacing="1", align_items="start", width="100%"
            ),
            padding="18px", bg=bg_color, border_radius="12px",
            border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW,
            width="100%", height="100%"
        )

    return rx.vstack(
        rx.grid(
            std_card("Total de dossiers", DashboardState.kpi_total_dossiers.to_string(), "Dossiers uniques", "folder"),
            std_card("Attente numérisation", DashboardState.kpi_attente_num.to_string(), "Flux restant", "scan"),
            std_card("Pièces indexées", DashboardState.kpi_pieces_indexees.to_string(), "Feuillets indexés", "file-text"),
            dark_card("Taux de restitution", f"{DashboardState.kpi_taux_restitution} %", "Dossiers restitués", "refresh-cw"),
            std_card("Durée moyenne scan", f"{DashboardState.kpi_duree_moyenne} h", "Délai indexation → retour", "clock"),
            std_card("Dossiers par jour", f"{DashboardState.kpi_dossiers_jour} /j", "Cadence moyenne", "trending-up"),
            std_card("Top archiviste", DashboardState.kpi_top_archiviste, f"Total : {DashboardState.kpi_top_total} dossiers", "user"),
            # v3.0 : Carte complexité moyenne globale
            complexite_card("Complexité moyenne", f"{DashboardState.kpi_complexite_moyenne}", "Propriétaires/dossier", "users", ""),
            columns="4", spacing="4", width="100%", margin_top="16px",
        ),
        # v3.0 : Ligne de complexité par bureau
        rx.cond(
            DashboardState.kpi_complexite_moyenne > 0,
            rx.grid(
                complexite_card("Complexité Cadastre", f"{DashboardState.kpi_complexite_cadastre}", "Propriétaires/dossier", "map-pin", "Cadastre"),
                complexite_card("Complexité Conservation", f"{DashboardState.kpi_complexite_conservation}", "Propriétaires/dossier", "shield", "Conservation"),
                complexite_card("Complexité Domaines", f"{DashboardState.kpi_complexite_domaines}", "Propriétaires/dossier", "globe", "Domaines"),
                columns="3", spacing="4", width="100%", margin_top="16px",
            ),
            rx.box()
        ),
        width="100%", spacing="0",
    )


# =============================================================================
# v2.8.7 — Conteneur agrandi pour éviter tout dépassement du SVG
# =============================================================================
def activity_chart() -> rx.Component:
    return rx.box(
        rx.vstack(
            rx.hstack(
                rx.hstack(
                    rx.icon("activity", size=16, color=GOLD),
                    rx.text("Flux d'activité", color=TEXT_MAIN, font_size="14px", font_weight="600"),
                    spacing="2", align_items="center"
                ),
                rx.spacer(),
                rx.hstack(
                    rx.button("Mois", on_click=DashboardState.set_flux_mois, bg=rx.cond(DashboardState.flux_period == "Mois", "#F0E6D0", WHITE), color=rx.cond(DashboardState.flux_period == "Mois", BROWN, TEXT_MUTED), border=f"1px solid {rx.cond(DashboardState.flux_period == 'Mois', GOLD, BORDER_CARD)}", border_radius="6px", padding="4px 12px", font_size="12px", height="32px", cursor="pointer"),
                    rx.button("Trimestre", on_click=DashboardState.set_flux_trimestre, bg=rx.cond(DashboardState.flux_period == "Trimestre", "#F0E6D0", WHITE), color=rx.cond(DashboardState.flux_period == "Trimestre", BROWN, TEXT_MUTED), border=f"1px solid {rx.cond(DashboardState.flux_period == 'Trimestre', GOLD, BORDER_CARD)}", border_radius="6px", padding="4px 12px", font_size="12px", height="32px", cursor="pointer"),
                    spacing="1",
                ),
                width="100%",
                align_items="center",
                padding="16px 20px 0 20px",
            ),
            rx.box(
                rx.html(DashboardState.flux_svg_html),
                width="100%",
                height="300px",
            ),
            width="100%",
            spacing="0",
        ),
        bg=WHITE,
        border_radius="12px",
        border=f"1px solid {BORDER_CARD}",
        box_shadow=SHADOW,
        width="100%",
        height="360px",
    )


def volume_donut() -> rx.Component:
    return rx.box(rx.vstack(
        rx.hstack(rx.icon("pie-chart", size=16, color=GOLD), rx.text("Volume par bureau", color=TEXT_MAIN, font_size="14px", font_weight="600"), spacing="2", align_items="center", padding="16px 20px 0 20px"),
        rx.html(DashboardState.donut_html), width="100%", spacing="0",
    ), bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%", height="380px")


def productivity_table() -> rx.Component:
    def agent_row(row: dict) -> rx.Component:
        return rx.hstack(
            rx.center(rx.text(row["rang"].to_string(), color=TEXT_MUTED, font_size="12px", font_weight="600"), width="32px", height="32px", border_radius="16px", bg=rx.cond(row["rang"] == 1, CREAM_CARD, "transparent")),
            rx.vstack(rx.text(row["agent"], color=TEXT_MAIN, font_size="13px", font_weight="600"), rx.text(row["csf"], color=TEXT_MUTED, font_size="11px"), spacing="0", align_items="start", width="160px"),
            rx.box(bureau_badge(row["bureau"]), width="110px"),
            rx.hstack(rx.box(rx.box(width=f"{row['pct']}%", height="6px", bg=GOLD, border_radius="3px"), width="120px", bg="#F0F0F0", border_radius="3px", height="6px"), rx.text(row["dossiers"].to_string(), color=TEXT_MAIN, font_size="13px", font_weight="600", width="50px", text_align="right"), spacing="2", align_items="center", flex="1"),
            rx.text(row["moyenne"].to_string(), color=TEXT_MAIN, font_size="13px", font_weight="500", width="60px", text_align="right"),
            width="100%", align_items="center", padding="10px 0", border_bottom=f"1px solid {BORDER_CARD}",
        )
    return rx.box(rx.vstack(
        rx.hstack(rx.hstack(rx.icon("users", size=16, color=GOLD), rx.text("Productivité équipe", color=TEXT_MAIN, font_size="14px", font_weight="600"), spacing="2", align_items="center"), rx.spacer(), rx.text(f"{DashboardState.productivite_count} agents", color=TEXT_MUTED, font_size="12px"), width="100%", align_items="center", padding="16px 20px"),
        rx.hstack(rx.text("AGENT", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", width="220px"), rx.text("BUREAU", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", width="110px"), rx.text("DOSSIERS TRAITÉS", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", flex="1"), rx.text("MOY./JOUR", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", width="70px", text_align="right"), width="100%", padding="0 20px 8px 20px", border_bottom=f"2px solid {BORDER_CARD}"),
        rx.vstack(rx.foreach(DashboardState.productivite, agent_row), width="100%", padding="0 20px", spacing="0"),
        width="100%", spacing="0",
    ), bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%")


def recent_records() -> rx.Component:
    def record_item(item: dict) -> rx.Component:
        return rx.hstack(
            rx.vstack(rx.text(item["code"], color=TEXT_MAIN, font_size="13px", font_weight="600"), rx.text(item["desc"], color=TEXT_MUTED, font_size="12px", max_width="280px"), spacing="1", align_items="start"),
            rx.spacer(),
            type_badge(item["type_badge"]),
            width="100%", align_items="center", padding="12px 0", border_bottom=f"1px solid {BORDER_CARD}",
        )
    return rx.box(rx.vstack(
        rx.hstack(rx.icon("file-text", size=16, color=GOLD), rx.text("Derniers enregistrements", color=TEXT_MAIN, font_size="14px", font_weight="600"), spacing="2", align_items="center", padding="16px 20px"),
        rx.vstack(rx.foreach(DashboardState.derniers_enregistrements, record_item), width="100%", padding="0 20px", spacing="0"),
        width="100%", spacing="0",
    ), bg=WHITE, border_radius="12px", border=f"1px solid {BORDER_CARD}", box_shadow=SHADOW, width="100%")


# v3.0 — NOUVEAU : Tableau de complexité par bureau
def complexity_table() -> rx.Component:
    def complexity_row(row: dict) -> rx.Component:
        return rx.hstack(
            rx.box(bureau_badge(row["bureau"]), width="120px"),
            rx.hstack(
                rx.box(
                    rx.box(
                        width=row["bar_width_str"],  # ← string pré-calculée
                        height="8px",
                        bg=rx.cond(row["is_conservation"], BROWN, GOLD),
                        border_radius="4px"
                    ),
                    width="100%",
                    bg="#F0F0F0",
                    border_radius="4px",
                    height="8px"
                ),
                rx.text(row["complexite"], color=TEXT_MAIN, font_size="14px", font_weight="600", width="50px", text_align="right"),
                spacing="2",
                align_items="center",
                flex="1"
            ),
            rx.text(
                row["categorie"],  # ← string pré-calculée
                color=rx.cond(
                    row["is_elevée"], RED_TEXT,
                    rx.cond(row["is_moyenne"], GOLD, GREEN_TEXT)
                ),
                font_size="12px",
                font_weight="500",
                width="80px",
                text_align="right"
            ),
            width="100%",
            align_items="center",
            padding="10px 0",
            border_bottom=f"1px solid {BORDER_CARD}",
        )
    return rx.box(
        rx.vstack(
            rx.hstack(
                rx.hstack(
                    rx.icon("bar-chart-3", size=16, color=GOLD),
                    rx.text("Complexité par bureau", color=TEXT_MAIN, font_size="14px", font_weight="600"),
                    spacing="2",
                    align_items="center"
                ),
                rx.spacer(),
                rx.text("Propriétaires uniques / dossier", color=TEXT_MUTED, font_size="11px"),
                width="100%",
                align_items="center",
                padding="16px 20px"
            ),
            rx.hstack(
                rx.text("BUREAU", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", width="120px"),
                rx.text("NIVEAU DE COMPLEXITÉ", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", flex="1"),
                rx.text("CATÉGORIE", color=TEXT_MUTED, font_size="10px", font_weight="700", letter_spacing="0.5px", width="80px", text_align="right"),
                width="100%",
                padding="0 20px 8px 20px",
                border_bottom=f"2px solid {BORDER_CARD}"
            ),
            rx.vstack(
                rx.foreach(DashboardState.complexite_bureau, complexity_row),
                width="100%",
                padding="0 20px",
                spacing="0"
            ),
            width="100%",
            spacing="0",
        ),
        bg=WHITE,
        border_radius="12px",
        border=f"1px solid {BORDER_CARD}",
        box_shadow=SHADOW,
        width="100%"
    )


# v3.0 — NOUVEAU : Stats Entrée/Sortie
def es_stats_section() -> rx.Component:
    return rx.box(
        rx.vstack(
            rx.hstack(
                rx.hstack(
                    rx.icon("arrow-left-right", size=16, color=GOLD),
                    rx.text("Balance Entrées / Sorties", color=TEXT_MAIN, font_size="14px", font_weight="600"),
                    spacing="2",
                    align_items="center"
                ),
                rx.spacer(),
                width="100%",
                align_items="center",
                padding="16px 20px"
            ),
            rx.hstack(
                rx.box(
                    rx.vstack(
                        rx.text("ENTRÉES", color=GREEN_TEXT, font_size="10px", font_weight="700", letter_spacing="1px"),
                        rx.text(DashboardState.stats_es["entrees"].to_string(), font_size="28px", font_weight="bold", color=GREEN_TEXT, line_height="1"),
                        rx.text("Enregistrements en entrée", color=TEXT_MUTED, font_size="11px"),
                        spacing="1",
                        align_items="center",
                        width="100%"
                    ),
                    padding="20px",
                    bg=GREEN_LIGHT,
                    border_radius="10px",
                    width="50%"
                ),
                rx.box(
                    rx.vstack(
                        rx.text("SORTIES", color=RED_TEXT, font_size="10px", font_weight="700", letter_spacing="1px"),
                        rx.text(DashboardState.stats_es["sorties"].to_string(), font_size="28px", font_weight="bold", color=RED_TEXT, line_height="1"),
                        rx.text("Enregistrements en sortie", color=TEXT_MUTED, font_size="11px"),
                        spacing="1",
                        align_items="center",
                        width="100%"
                    ),
                    padding="20px",
                    bg=RED_LIGHT,
                    border_radius="10px",
                    width="50%"
                ),
                spacing="4",
                width="100%",
                padding="0 20px 20px 20px"
            ),
            width="100%",
            spacing="0",
        ),
        bg=WHITE,
        border_radius="12px",
        border=f"1px solid {BORDER_CARD}",
        box_shadow=SHADOW,
        width="100%"
    )


def loading_spinner() -> rx.Component:
    return rx.center(rx.vstack(rx.spinner(size="3", color=GOLD, thickness="4px"), rx.text("Chargement des données...", color=TEXT_MUTED, font_size="14px", margin_top="16px")), width="100%", height="100vh", bg=CREAM)


def error_banner() -> rx.Component:
    return rx.cond(DashboardState.db_error != "", rx.box(rx.hstack(rx.icon("triangle_alert", size=16, color=RED_TEXT), rx.text(DashboardState.db_error, color=RED_TEXT, font_size="13px"), spacing="2", align_items="center"), padding="12px 24px", bg=RED_LIGHT, border_radius="8px", margin="16px 32px", width="calc(100% - 64px)"), rx.box())


def index() -> rx.Component:
    return rx.box(
        rx.cond(DashboardState.is_loading & (DashboardState.kpi_total_dossiers == 0), loading_spinner(),
            rx.vstack(
                header(),
                filter_bar(),
                error_banner(),
                rx.box(
                    rx.vstack(
                        objectives_section(),
                        metrics_section(),
                        rx.hstack(
                            rx.box(activity_chart(), width="65%"),
                            rx.box(volume_donut(), width="35%"),
                            spacing="4",
                            width="100%",
                            align_items="stretch"
                        ),
                        # v3.0 : Nouvelle ligne avec stats E/S et complexité
                        rx.hstack(
                            rx.box(es_stats_section(), width="40%"),
                            rx.box(complexity_table(), width="60%"),
                            spacing="4",
                            width="100%",
                            align_items="stretch"
                        ),
                        rx.hstack(
                            rx.box(productivity_table(), width="60%"),
                            rx.box(recent_records(), width="40%"),
                            spacing="4",
                            width="100%",
                            align_items="stretch"
                        ),
                        spacing="6",
                        width="100%",
                        max_width="1400px",
                        margin="0 auto",
                        padding="24px 0",
                    ),
                    width="100%",
                    bg=CREAM,
                    min_height="100vh"
                ),
                spacing="0",
                width="100%"
            ),
        ),
        width="100%",
        min_height="100vh",
        bg=CREAM,
        on_mount=DashboardState.on_load,
    )


app = rx.App(theme=rx.theme(appearance="light", accent_color="orange", radius="medium"), stylesheets=[])
app.add_page(index, route="/", title="Dashboard de Suivi — DGID")